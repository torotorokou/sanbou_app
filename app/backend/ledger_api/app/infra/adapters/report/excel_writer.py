from openpyxl import load_workbook
from io import BytesIO
import pandas as pd
import numpy as np
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from copy import copy
from pathlib import Path
import os
import unicodedata
from typing import Any, cast
import re
from app.infra.report_utils import app_logger


def safe_excel_value(value):
    """Excelに書き込める形式に変換するユーティリティ関数"""
    if pd.isna(value) or value is pd.NA or value is np.nan:
        return None
    elif isinstance(value, (dict, list, set)):
        return str(value)
    elif hasattr(value, "strftime"):
        return value.strftime("%Y/%m/%d")
    return value


def load_template_workbook(template_path: str | Path) -> Workbook:
    """
    テンプレートExcelを読み込む。
    見つからない/読み込めない場合は空のWorkbookでフォールバックし、サーバーログに警告を出す。
    """
    logger = app_logger()
    # BASE_API_DIR環境変数を使用、未設定の場合はapiディレクトリを基準とする
    current_file_dir = Path(__file__).parent.parent  # utils -> ledger
    base_dir = Path(os.getenv("BASE_API_DIR", str(current_file_dir.parent.parent.parent.parent.parent / "api")))
    full_path = base_dir / Path(template_path)

    try:
        if not full_path.exists():
            logger.warning(
                f"テンプレートExcelが見つかりませんでした。パス: {full_path}. 空のWorkbookで代替します。"
            )
            return Workbook()
        return load_workbook(full_path)
    except Exception as e:
        logger.warning(
            f"テンプレートExcelの読み込みに失敗しました。パス: {full_path}. 空のWorkbookで代替します。理由: {e}"
        )
        return Workbook()


def _resolve_fallback_font_name(font_name: str | None) -> str | None:
    if not font_name:
        return None

    normalized = unicodedata.normalize("NFKC", font_name).strip()
    compact = normalized.replace(" ", "")

    fallback_map = {
        "MSPゴシック": "Noto Sans CJK JP",
        "MSPGothic": "Noto Sans CJK JP",
        "MSGothic": "Noto Sans CJK JP",
        "MSゴシック": "Noto Sans CJK JP",
        "MS明朝": "Noto Sans CJK JP",
        "游ゴシック": "Noto Sans CJK JP",
        "YuGothic": "Noto Sans CJK JP",
        "YuGothicUI": "Noto Sans CJK JP",
    }

    return fallback_map.get(compact)


def _maybe_replace_font(font: Any, tracker: set[tuple[str, str]]) -> Any:
    name = getattr(font, "name", None)
    target = _resolve_fallback_font_name(name)
    if target and target != name:
        tracker.add(((name or ""), target))
        try:
            cloned = copy(font)
            if hasattr(cloned, "name"):
                object.__setattr__(cloned, "name", target)
            return cloned
        except Exception:
            if hasattr(font, "copy") and callable(getattr(font, "copy")):
                return font.copy(name=target)  # type: ignore[attr-defined]
    return font


def normalize_workbook_fonts(wb: Workbook, logger=None) -> None:
    tracker: set[tuple[str, str]] = set()

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                font = getattr(cell, "font", None)
                if font is None:
                    continue
                new_font = _maybe_replace_font(font, tracker)
                if new_font is not font:
                    try:
                        cell.font = new_font
                    except Exception:
                        continue

    fonts_container = getattr(wb, "_fonts", None)
    if fonts_container:
        for idx, font in enumerate(list(fonts_container)):
            if font is None:
                continue
            new_font = _maybe_replace_font(font, tracker)
            if new_font is not font:
                try:
                    fonts_container[idx] = new_font
                except Exception:
                    continue

    named_styles = getattr(wb, "_named_styles", None)
    if named_styles:
        for style in named_styles:
            font = getattr(style, "font", None)
            if font is None:
                continue
            new_font = _maybe_replace_font(font, tracker)
            if new_font is not font:
                try:
                    style.font = new_font
                except Exception:
                    continue

    if tracker:
        logger = logger or app_logger()
        replacements = ", ".join(f"{src}→{dst}" for src, dst in sorted(tracker))
        logger.info(f"Excelテンプレートのフォントを標準フォントに置換しました: {replacements}")


def write_dataframe_to_worksheet(df: pd.DataFrame, ws: Worksheet, logger=None):
    if logger is None:
        logger = app_logger()

    for idx, row in df.iterrows():
        cell_ref = row.get("セル")
        value = safe_excel_value(row.get("値"))

        # セル参照の基本的なバリデーション（A1形式）
        cell_ref_str = str(cell_ref).strip() if not pd.isna(cell_ref) else ""
        if (
            pd.isna(cell_ref)
            or cell_ref_str in ["", "未設定"]
            or not re.match(r"^[A-Za-z]+[0-9]+$", cell_ref_str)
        ):
            logger.info(f"空欄または未設定のセルはスキップされました。行 {idx}")
            continue

        try:
            cell = cast(Cell, ws[cell_ref_str])

            if isinstance(cell, MergedCell):
                logger.warning(f"セル {cell_ref} は結合セルで書き込み不可。値: {value}")
                continue

            # --- 書式をdeep copyで保持 ---
            original_font = copy(cell.font)
            original_fill = copy(cell.fill)
            original_border = copy(cell.border)
            original_format = cell.number_format

            # 値の上書き
            cell.value = value

            # --- 書式の復元 ---
            cell.font = original_font  # type: ignore[assignment]
            cell.fill = original_fill  # type: ignore[assignment]
            cell.border = original_border  # type: ignore[assignment]
            cell.number_format = original_format

        except Exception as e:
            logger.error(f"セル {cell_ref} 書き込み失敗: {e} / 値: {value}")


def rename_sheet(wb: Workbook, new_title: str):
    # Excelシート名で使用できない文字を置換
    invalid_chars = [":", "/", "\\", "?", "*", "[", "]"]
    sanitized_title = new_title
    for char in invalid_chars:
        sanitized_title = sanitized_title.replace(char, "_")

    # シート名の長さ制限（Excelは31文字まで）
    if len(sanitized_title) > 31:
        sanitized_title = sanitized_title[:31]

    ws = wb.active
    ws.title = sanitized_title  # type: ignore[attr-defined]


def save_workbook_to_bytesio(wb: Workbook) -> BytesIO:
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def write_values_to_template(
    df: pd.DataFrame, template_path: str, extracted_date: str
) -> BytesIO:
    """
    単一責任原則に基づいて分割されたExcelテンプレート書き込み関数
    - テンプレ読み込み
    - セルへの書き込み
    - シート名変更
    - メモリ出力
    """
    logger = app_logger()
    wb = load_template_workbook(template_path)
    normalize_workbook_fonts(wb, logger=logger)
    ws = wb.active
    if not isinstance(ws, Worksheet):
        raise TypeError("Workbook のアクティブシートが Worksheet ではありません")

    write_dataframe_to_worksheet(df, ws, logger=logger)
    rename_sheet(wb, extracted_date)
    return save_workbook_to_bytesio(wb)
