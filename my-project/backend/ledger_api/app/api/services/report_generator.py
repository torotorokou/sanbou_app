# backend/app/api/services/report_generator.py

import os
import pandas as pd
from datetime import datetime
from abc import ABC, abstractmethod
from typing import Dict, Any, Optional
from pathlib import Path

from backend_shared.config.config_loader import ReportTemplateConfigLoader
from backend_shared.src.report_checker.check_csv_files import check_csv_files

# 先ほど作成したExcel出力機能をインポート
from app.api.services.utils.make_excel import df_to_excel

# reportlabがインストールされている場合のみPDF機能を有効化
try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# factory_report.pyのprocess関数を直接呼び出す
from .manage_report_processors.factory_report.main import (
    factory_report_main_process,
)


class BaseReportGenerator(ABC):
    """
    レポート生成の基底クラス

    全ての帳票生成クラスの共通機能を提供します：
    - 前処理（ファイルチェック）
    - メイン処理（サブクラスで実装）
    - PDF生成
    - Excel生成
    - 後処理（ファイル名生成）
    """

    def __init__(self, output_dir: str, files: Dict[str, Any]):
        """
        Args:
            output_dir: 出力ディレクトリのパス
            files: CSVファイルのDataFrame辞書 {ファイル種別: DataFrame}
        """
        self.output_dir = output_dir
        self.files = files
        self.config_loader_report = ReportTemplateConfigLoader()
        self.result_df = None  # main_processの結果を保存

    def preprocess(self, report_key: Optional[str] = None):
        """
        前処理：必要なファイルがそろっているかチェック

        Args:
            report_key: レポートの種類（例: "factory_report"）
        """
        if report_key is None:
            raise ValueError("report_key must be provided and must be a string.")

        # 設定ファイルから必要なファイル一覧を取得
        required = self.config_loader_report.get_required_files(report_key)
        optional = self.config_loader_report.get_optional_files(report_key)

        # ファイルチェック実行
        check_csv_files(self.files, required, optional)
        print(f"[INFO] 必須ファイルチェックOK: {required} がすべて揃っています。")
        return self.files

    @abstractmethod
    def main_process(self) -> pd.DataFrame:
        """
        メインの帳票生成処理（サブクラスで実装必須）

        Returns:
            pd.DataFrame: 生成された帳票データ
        """
        raise NotImplementedError("Subclasses must implement this method.")

    def generate_pdf(self, pdf_name: str) -> str:
        """
        エクセルファイルを作成してからPDFに変換

        Args:
            pdf_name: 生成するPDFファイル名

        Returns:
            str: 生成されたPDFファイル名
        """
        # 出力ディレクトリが存在しない場合は作成
        os.makedirs(self.output_dir, exist_ok=True)
        pdf_path = os.path.join(self.output_dir, pdf_name)

        if self.result_df is not None and isinstance(self.result_df, pd.DataFrame):
            print(f"[INFO] DataFrameからExcel経由でPDF生成: {len(self.result_df)}行")

            # 1. まずExcelファイルを作成
            excel_name = pdf_name.replace(".pdf", ".xlsx")
            excel_path = self.generate_excel(excel_name)
            excel_full_path = os.path.join(self.output_dir, excel_path)

            # 2. ExcelファイルをPDFに変換
            self._convert_excel_to_pdf(excel_full_path, pdf_path)

        else:
            print("[ERROR] result_dfがありません。エラー情報PDFを生成します。")
            self._generate_error_pdf(pdf_path)

        print(f"[INFO] PDF生成完了: {pdf_path}")
        return pdf_name

    def _convert_excel_to_pdf(self, excel_path: str, pdf_path: str):
        """
        ExcelファイルをPDFに変換

        Args:
            excel_path: 変換元のExcelファイルパス
            pdf_path: 出力するPDFファイルパス
        """
        try:
            if REPORTLAB_AVAILABLE:
                self._convert_excel_to_pdf_reportlab(excel_path, pdf_path)
            else:
                self._convert_excel_to_pdf_basic(excel_path, pdf_path)
        except Exception as e:
            print(f"[ERROR] Excel→PDF変換エラー: {e}")
            # エラー時は基本的なPDFを生成
            self._generate_error_pdf(pdf_path)

    def _convert_excel_to_pdf_reportlab(self, excel_path: str, pdf_path: str):
        """
        reportlabを使用してExcelファイルをPDFに変換
        """
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        import openpyxl

        try:
            # Excelファイルを読み込み
            workbook = openpyxl.load_workbook(excel_path)
            worksheet = workbook.active

            # PDF文書を作成
            doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            elements = []

            # タイトル
            title = Paragraph(
                f"帳票レポート（Excel変換） - {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}",
                styles["Title"],
            )
            elements.append(title)
            elements.append(Spacer(1, 0.2 * inch))

            # worksheetがNoneでないことを確認
            if worksheet is not None:
                # Excelのデータをテーブルデータに変換
                table_data = []
                max_rows = min(100, worksheet.max_row or 1)  # 最大100行まで
                max_cols = min(20, worksheet.max_column or 1)  # 最大20列まで

                for row in range(1, max_rows + 1):
                    row_data = []
                    for col in range(1, max_cols + 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value is None:
                            cell_value = ""
                        else:
                            cell_value = str(cell_value)
                            # 長すぎる文字列は切り詰め
                            if len(cell_value) > 15:
                                cell_value = cell_value[:12] + "..."
                        row_data.append(cell_value)
                    table_data.append(row_data)

                if table_data:
                    # テーブルを作成
                    table = Table(table_data)

                    # テーブルスタイルを適用
                    style_commands = [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 1), (-1, -1), 7),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]

                    table.setStyle(TableStyle(style_commands))
                    elements.append(table)

                    # 注意書き
                    if (worksheet.max_row or 0) > 100 or (
                        worksheet.max_column or 0
                    ) > 20:
                        elements.append(Spacer(1, 0.2 * inch))
                        note = Paragraph(
                            f"注意: 表示制限により一部のみ表示（全{worksheet.max_row or 0}行×{worksheet.max_column or 0}列）",
                            styles["Italic"],
                        )
                        elements.append(note)
                else:
                    # データがない場合
                    no_data = Paragraph(
                        "データが見つかりませんでした", styles["Normal"]
                    )
                    elements.append(no_data)
            else:
                # worksheetがない場合
                error_msg = Paragraph(
                    "Excelファイルの読み込みに失敗しました", styles["Normal"]
                )
                elements.append(error_msg)

            # PDFを生成
            doc.build(elements)
            workbook.close()
            print("[INFO] reportlabを使用してExcel→PDF変換完了")

        except Exception as e:
            print(f"[ERROR] Excel→PDF変換中にエラー: {e}")
            # エラー時は基本的なエラーPDFを生成
            self._generate_error_pdf(pdf_path)

    def _convert_excel_to_pdf_basic(self, excel_path: str, pdf_path: str):
        """
        基本的なExcel→PDF変換（reportlabなし）
        """
        import openpyxl

        try:
            # Excelファイルを読み込み
            workbook = openpyxl.load_workbook(excel_path)
            worksheet = workbook.active

            # 基本情報を取得
            if worksheet is not None:
                row_count = worksheet.max_row or 0
                col_count = worksheet.max_column or 0

                # サンプルデータを取得（最初の5行5列）
                sample_data = []
                for row in range(1, min(6, row_count + 1)):
                    row_data = []
                    for col in range(1, min(6, col_count + 1)):
                        cell_value = worksheet.cell(row=row, column=col).value
                        row_data.append(
                            str(cell_value) if cell_value is not None else ""
                        )
                    sample_data.append(" | ".join(row_data))
            else:
                row_count = 0
                col_count = 0
                sample_data = ["データなし"]

            # 基本的なPDF内容を生成
            pdf_content = f"""%PDF-1.4
1 0 obj
<<
/Type /Catalog
/Pages 2 0 R
>>
endobj

2 0 obj
<<
/Type /Pages
/Kids [3 0 R]
/Count 1
>>
endobj

3 0 obj
<<
/Type /Page
/Parent 2 0 R
/MediaBox [0 0 612 792]
/Contents 4 0 R
/Resources <<
/Font <<
/F1 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>
>>
>>
>>
endobj

4 0 obj
<<
/Length 1000
>>
stream
BT
/F1 16 Tf
50 750 Td
(Excel to PDF Report - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}) Tj
0 -30 Td
/F1 12 Tf
(Source Excel: {os.path.basename(excel_path)}) Tj
0 -20 Td
(Total Rows: {row_count}) Tj
0 -20 Td
(Total Columns: {col_count}) Tj
0 -30 Td
(Sample Data:) Tj
0 -20 Td
/F1 10 Tf
({chr(10).join(sample_data[:3]).replace(chr(10), " ")}) Tj
ET
endstream
endobj

xref
0 5
0000000000 65535 f 
0000000010 00000 n 
0000000053 00000 n 
0000000100 00000 n 
0000000300 00000 n 
trailer
<<
/Size 5
/Root 1 0 R
>>
startxref
1500
%%EOF""".encode("utf-8")

            with open(pdf_path, "wb") as f:
                f.write(pdf_content)

            workbook.close()
            print("[INFO] 基本的なExcel→PDF変換完了")

        except Exception as e:
            print(f"[ERROR] 基本Excel→PDF変換中にエラー: {e}")
            # エラー時は基本的なエラーPDFを生成
            self._generate_error_pdf(pdf_path)

    def _generate_reportlab_pdf(self, pdf_path: str):
        """reportlabを使用した高品質PDF生成"""
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.units import inch

        # 型安全性のためのアサーション
        assert self.result_df is not None, "result_df should not be None"
        df = self.result_df  # 型チェック用の変数

        # PDF文書を作成
        doc = SimpleDocTemplate(pdf_path, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # タイトル
        title = Paragraph(
            f"帳票レポート - {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}",
            styles["Title"],
        )
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        # 概要情報
        summary = Paragraph(
            f"総レコード数: {len(df)}行<br/>総カラム数: {len(df.columns)}列",
            styles["Normal"],
        )
        elements.append(summary)
        elements.append(Spacer(1, 0.3 * inch))

        # データテーブル（最大20行まで表示）
        display_df = df.head(20) if len(df) > 20 else df

        # テーブルデータを準備
        table_data = [list(display_df.columns)]  # ヘッダー
        for _, row in display_df.iterrows():
            # 各セルの値を文字列に変換し、長すぎる場合は切り詰め
            row_data = []
            for value in row:
                str_value = str(value)
                if len(str_value) > 20:
                    str_value = str_value[:17] + "..."
                row_data.append(str_value)
            table_data.append(row_data)

        # テーブルを作成
        table = Table(table_data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(table)

        if len(df) > 20:
            elements.append(Spacer(1, 0.2 * inch))
            note = Paragraph(
                f"注意: 表示は最初の20行のみです。全{len(df)}行はExcelファイルで確認できます。",
                styles["Italic"],
            )
            elements.append(note)

        # PDFを生成
        doc.build(elements)
        print("[INFO] 高品質PDF生成完了（reportlab使用）")

    def _generate_basic_pdf(self, pdf_path: str):
        """基本的なPDF生成（reportlabなし）"""
        # 型安全性のためのアサーション
        assert self.result_df is not None, "result_df should not be None"
        df = self.result_df  # 型チェック用の変数

        # DataFrameの基本情報を取得
        row_count = len(df)
        col_count = len(df.columns)
        columns = list(df.columns)

        # DataFrameの内容をテキストとして取得（最初の5行）
        df_sample = df.head(5).to_string(index=False, max_cols=10)

        # 実際のデータを含むPDF内容を生成
        pdf_content = f"""%PDF-1.4
1 0 obj
<<
/Type /Catalog
/Pages 2 0 R
>>
endobj

2 0 obj
<<
/Type /Pages
/Kids [3 0 R]
/Count 1
>>
endobj

3 0 obj
<<
/Type /Page
/Parent 2 0 R
/MediaBox [0 0 612 792]
/Contents 4 0 R
/Resources <<
/Font <<
/F1 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>
>>
>>
>>
endobj

4 0 obj
<<
/Length 800
>>
stream
BT
/F1 16 Tf
50 750 Td
(Generated Report - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}) Tj
0 -30 Td
/F1 12 Tf
(Total Records: {row_count}) Tj
0 -20 Td
(Total Columns: {col_count}) Tj
0 -20 Td
(Columns: {", ".join(columns[:5])}) Tj
0 -30 Td
(Data Sample:) Tj
0 -20 Td
/F1 10 Tf
({df_sample.replace(chr(10), " ")[:200]}...) Tj
ET
endstream
endobj

xref
0 5
0000000000 65535 f 
0000000010 00000 n 
0000000053 00000 n 
0000000100 00000 n 
0000000300 00000 n 
trailer
<<
/Size 5
/Root 1 0 R
>>
startxref
1200
%%EOF""".encode("utf-8")

        with open(pdf_path, "wb") as f:
            f.write(pdf_content)
        print("[INFO] 基本PDF生成完了")

    def _generate_error_pdf(self, pdf_path: str):
        """エラー情報PDF生成"""
        pdf_content = f"""%PDF-1.4
1 0 obj
<<
/Type /Catalog
/Pages 2 0 R
>>
endobj

2 0 obj
<<
/Type /Pages
/Kids [3 0 R]
/Count 1
>>
endobj

3 0 obj
<<
/Type /Page
/Parent 2 0 R
/MediaBox [0 0 612 792]
/Contents 4 0 R
/Resources <<
/Font <<
/F1 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>
>>
>>
>>
endobj

4 0 obj
<<
/Length 200
>>
stream
BT
/F1 16 Tf
50 750 Td
(Report Generation Error) Tj
0 -30 Td
/F1 12 Tf
(No data available for report generation) Tj
0 -20 Td
(Generated at: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}) Tj
ET
endstream
endobj

xref
0 5
0000000000 65535 f 
0000000010 00000 n 
0000000053 00000 n 
0000000100 00000 n 
0000000300 00000 n 
trailer
<<
/Size 5
/Root 1 0 R
>>
startxref
600
%%EOF""".encode("utf-8")

        with open(pdf_path, "wb") as f:
            f.write(pdf_content)
        print("[WARNING] エラー情報PDF生成完了")

    def generate_excel(self, excel_name: str) -> str:
        """
        Excelファイルを生成（テンプレートを使用してDataFrameデータを埋め込み）

        Args:
            excel_name: 生成するExcelファイル名

        Returns:
            str: 生成されたExcelファイル名
        """
        # 出力ディレクトリが存在しない場合は作成
        os.makedirs(self.output_dir, exist_ok=True)
        excel_path = os.path.join(self.output_dir, excel_name)

        if self.result_df is not None and isinstance(self.result_df, pd.DataFrame):
            # 実際のDataFrameを使用してExcel生成
            print(
                f"[INFO] 実際のDataFrameを使用してExcel生成: {len(self.result_df)}行, {len(self.result_df.columns)}列"
            )
            print(f"[INFO] 列名: {list(self.result_df.columns)}")

            # テンプレートパスを取得（サブクラスで実装される可能性を考慮）
            template_path = self._get_template_path()

            if template_path and os.path.exists(template_path):
                print(f"[INFO] テンプレートファイルを使用: {template_path}")
                # テンプレート機能を使用してExcel生成
                excel_bytes = df_to_excel(
                    df=self.result_df,
                    sheet_name="帳票データ",
                    use_formatting=True,
                    template_path=template_path,  # テンプレートパスを指定
                )
            else:
                print(
                    "[INFO] テンプレートファイルが見つからないため、標準フォーマットを使用"
                )
                # 標準のフォーマット機能を使用
                excel_bytes = df_to_excel(
                    df=self.result_df,
                    sheet_name="帳票データ",
                    use_formatting=True,  # 日本語フォント付きフォーマットを使用
                )

            # ファイルに保存
            with open(excel_path, "wb") as f:
                f.write(excel_bytes)

            print(
                f"[INFO] 実際のデータでExcel生成成功: {len(self.result_df)}行のデータを出力"
            )

        else:
            # DataFrameがない場合はエラー情報を含むExcel生成
            print("[ERROR] result_dfがありません。エラー情報Excelを生成します。")
            error_df = pd.DataFrame(
                {
                    "エラー": ["データが生成されませんでした"],
                    "詳細": ["main_process()でDataFrameが作成されていません"],
                    "作成日時": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                    "対処方法": ["main_process()の実装を確認してください"],
                }
            )

            excel_bytes = df_to_excel(
                df=error_df, sheet_name="エラー情報", use_formatting=True
            )

            with open(excel_path, "wb") as f:
                f.write(excel_bytes)

            print("[WARNING] エラー情報でExcel生成完了")

        print(f"[INFO] Excel生成完了: {excel_path}")
        return excel_name

    def _get_template_path(self) -> Optional[str]:
        """
        テンプレートファイルのパスを取得（サブクラスでオーバーライド可能）

        Returns:
            Optional[str]: テンプレートファイルのパス、存在しない場合はNone
        """
        return None  # 基底クラスではテンプレートなし

    def postprocess(self) -> str:
        """後処理：ダウンロード名用の日付返却"""
        return datetime.now().strftime("%Y%m%d_%H%M%S")

    def get_download_pdf_name(self, file_name_jp: str, date_str: str) -> str:
        """PDF用のダウンロードファイル名を生成"""
        return f"{file_name_jp}_{date_str}.pdf"

    def get_download_excel_name(self, file_name_jp: str, date_str: str) -> str:
        """Excel用のダウンロードファイル名を生成"""
        return f"{file_name_jp}_{date_str}.xlsx"


class FactoryReportGenerator(BaseReportGenerator):
    """工場レポート生成クラス"""

    def main_process(self) -> pd.DataFrame:
        """
        工場レポートのメイン処理

        Returns:
            pd.DataFrame: 生成された工場レポートデータ
        """
        print("[INFO] 工場レポート生成開始")

        # factory_report_main_processを呼び出し
        result_df = factory_report_main_process(self.files)

        # 結果をインスタンス変数に保存（PDF/Excel生成で使用）
        self.result_df = result_df

        print(f"[INFO] 工場レポート生成完了: {len(result_df)}行")
        return result_df

    def _get_template_path(self) -> Optional[str]:
        """
        工場レポート用のテンプレートファイルパスを取得

        Returns:
            Optional[str]: テンプレートファイルのパス
        """
        try:
            # YAMLファイルからテンプレートパスを取得
            template_path = self.config_loader_report.get_template_excel_path(
                "factory_report"
            )

            if template_path and os.path.exists(template_path):
                print(f"[INFO] 工場レポートテンプレートファイル発見: {template_path}")
                return template_path
            else:
                print(
                    f"[WARNING] 工場レポートテンプレートファイルが存在しません: {template_path}"
                )
                return None

        except Exception as e:
            print(f"[ERROR] テンプレートパス取得エラー: {e}")
            return None


class BalanceSheetGenerator(BaseReportGenerator):
    """バランスシート生成クラス（未実装）"""

    def main_process(self) -> pd.DataFrame:
        """
        バランスシートのメイン処理（サンプル実装）

        Returns:
            pd.DataFrame: 生成されたバランスシートデータ
        """
        print("[INFO] バランスシート生成開始（サンプル実装）")

        # サンプルデータを作成
        sample_data = pd.DataFrame(
            {
                "項目": ["売上", "費用", "利益"],
                "金額": [1000000, 600000, 400000],
                "比率": ["100%", "60%", "40%"],
            }
        )

        self.result_df = sample_data
        print(f"[INFO] バランスシート生成完了（サンプル）: {len(sample_data)}行")
        return sample_data


class AverageSheetGenerator(BaseReportGenerator):
    """平均シート生成クラス（未実装）"""

    def main_process(self) -> pd.DataFrame:
        """
        平均シートのメイン処理（サンプル実装）

        Returns:
            pd.DataFrame: 生成された平均シートデータ
        """
        print("[INFO] 平均シート生成開始（サンプル実装）")

        # サンプルデータを作成
        sample_data = pd.DataFrame(
            {
                "期間": ["1月", "2月", "3月"],
                "平均値": [150.5, 175.2, 163.8],
                "最高値": [200, 220, 195],
                "最低値": [100, 130, 125],
            }
        )

        self.result_df = sample_data
        print(f"[INFO] 平均シート生成完了（サンプル）: {len(sample_data)}行")
        return sample_data


class BlockUnitPriceGenerator(BaseReportGenerator):
    """ブロック単価生成クラス（未実装）"""

    def main_process(self) -> pd.DataFrame:
        """
        ブロック単価のメイン処理（サンプル実装）

        Returns:
            pd.DataFrame: 生成されたブロック単価データ
        """
        print("[INFO] ブロック単価生成開始（サンプル実装）")

        # サンプルデータを作成
        sample_data = pd.DataFrame(
            {
                "ブロック名": ["Aブロック", "Bブロック", "Cブロック"],
                "単価": [1500, 1800, 1200],
                "数量": [10, 8, 15],
                "合計": [15000, 14400, 18000],
            }
        )

        self.result_df = sample_data
        print(f"[INFO] ブロック単価生成完了（サンプル）: {len(sample_data)}行")
        return sample_data


class ManagementSheetGenerator(BaseReportGenerator):
    """管理シート生成クラス（未実装）"""

    def main_process(self) -> pd.DataFrame:
        """
        管理シートのメイン処理（サンプル実装）

        Returns:
            pd.DataFrame: 生成された管理シートデータ
        """
        print("[INFO] 管理シート生成開始（サンプル実装）")

        # サンプルデータを作成
        sample_data = pd.DataFrame(
            {
                "管理項目": ["品質", "コスト", "納期"],
                "目標値": [95, 1000000, 30],
                "実績値": [97, 950000, 28],
                "達成率": ["102%", "105%", "107%"],
            }
        )

        self.result_df = sample_data
        print(f"[INFO] 管理シート生成完了（サンプル）: {len(sample_data)}行")
        return sample_data


class BalanceManagementTableGenerator(BaseReportGenerator):
    """残高管理表生成クラス（未実装）"""

    def main_process(self) -> pd.DataFrame:
        """
        残高管理表のメイン処理（サンプル実装）

        Returns:
            pd.DataFrame: 生成された残高管理表データ
        """
        print("[INFO] 残高管理表生成開始（サンプル実装）")

        # サンプルデータを作成
        sample_data = pd.DataFrame(
            {
                "日付": ["2025-07-01", "2025-07-15", "2025-07-31"],
                "入金": [500000, 300000, 800000],
                "出金": [200000, 450000, 100000],
                "残高": [300000, 150000, 850000],
            }
        )

        self.result_df = sample_data
        print(f"[INFO] 残高管理表生成完了（サンプル）: {len(sample_data)}行")
        return sample_data


def get_report_generator(
    report_key: str, output_dir: str, files: Dict[str, Any]
) -> BaseReportGenerator:
    """
    帳簿種別に応じて適切なGeneratorを返す

    Args:
        report_key: レポートの種類（"factory_report", "balance_sheet"など）
        output_dir: 出力ディレクトリのパス
        files: CSVファイルのDataFrame辞書

    Returns:
        BaseReportGenerator: 対応するレポート生成クラスのインスタンス

    Example:
        # エンドポイントでの使用例
        generator = get_report_generator("factory_report", "/output", df_dict)
        generator.preprocess("factory_report")
        result_df = generator.main_process()
        pdf_name = generator.generate_pdf("report.pdf")
        excel_name = generator.generate_excel("report.xlsx")
    """
    generators = {
        "factory_report": FactoryReportGenerator,
        "balance_sheet": BalanceSheetGenerator,
        "average_sheet": AverageSheetGenerator,
        "block_unit_price": BlockUnitPriceGenerator,
        "management_sheet": ManagementSheetGenerator,
        "balance_management_table": BalanceManagementTableGenerator,
    }

    if report_key not in generators:
        available_keys = list(generators.keys())
        raise ValueError(
            f"Unsupported report type: {report_key}. Available types: {available_keys}"
        )

    generator_class = generators[report_key]
    return generator_class(output_dir, files)
