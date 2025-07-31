"""
Excel出力ユーティリティ（FastAPI用）

DataFrameをExcelファイルとして出力するための機能を提供します。
- シンプルなDataFrame→Excel変換
- テンプレートベースの変換
- 日本語フォント対応
"""

import io
import os
import pandas as pd
from io import BytesIO
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Optional, Union, Any
import logging

logger = logging.getLogger(__name__)


def safe_excel_value(value: Any) -> Any:
    """Excelに書き込める形式に変換するユーティリティ関数"""
    if pd.isna(value) or value is pd.NA or value is np.nan:
        return None
    elif isinstance(value, (dict, list, set)):
        return str(value)
    elif hasattr(value, "strftime"):
        return value.strftime("%Y/%m/%d")
    return value


def simple_df_to_excel(df: pd.DataFrame, output_path: str, sheet_name: str = "Sheet1"):
    """
    DataFrameを単純なExcelファイルとして保存

    Args:
        df: 保存するDataFrame
        output_path: 出力先パス
        sheet_name: シート名
    """
    # NaN値を適切に処理
    df_clean = df.fillna("")  # NaN値を空文字列に置換

    # 出力ディレクトリが存在しない場合は作成
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # xlsxwriterの代わりにopenpyxlを使用してNaN処理を回避
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_clean.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"[INFO] 単純Excel生成完了: {output_path}")
    except Exception as e:
        print(f"[ERROR] Excel生成エラー: {e}")
        # フォールバック: CSVとして保存
        csv_path = output_path.replace(".xlsx", ".csv")
        df_clean.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"[INFO] フォールバックCSV生成: {csv_path}")


def formatted_df_to_excel(df: pd.DataFrame, sheet_name: str = "データ") -> bytes:
    """
    フォーマット付きDataFrame→Excel変換（日本語フォント対応）

    Args:
        df: 出力するDataFrame
        sheet_name: シート名

    Returns:
        bytes: Excelファイルのバイナリデータ
    """
    output = BytesIO()

    # NaNや文字列'nan'などを空白に変換（中項目のみ）
    df_clean = df.copy()

    # NaN、INF、None値の包括的な処理
    print(f"[DEBUG] Excel変換前のDataFrame情報:")
    print(f"  - 行数: {len(df_clean)}, 列数: {len(df_clean.columns)}")
    print(f"  - 列名: {list(df_clean.columns)}")

    # 数値列のNaN/INF処理
    for col in df_clean.columns:
        if df_clean[col].dtype in ["float64", "float32", "int64", "int32"]:
            # NaN値を0に置換
            nan_count = df_clean[col].isna().sum()
            if nan_count > 0:
                print(f"  - 列 '{col}': {nan_count}個のNaN値を0に置換")
                df_clean[col] = df_clean[col].fillna(0)

            # INF値を0に置換
            import numpy as np

            inf_count = np.isinf(df_clean[col]).sum()
            if inf_count > 0:
                print(f"  - 列 '{col}': {inf_count}個のINF値を0に置換")
                df_clean[col] = df_clean[col].replace([np.inf, -np.inf], 0)

    # 文字列列のNaN処理
    if "中項目" in df_clean.columns:
        df_clean["中項目"] = (
            df_clean["中項目"]
            .replace(["nan", "NaN", "None"], "")
            .fillna("")
            .astype(str)
        )

    # 全ての列で残ったNaN値を適切な値に置換
    for col in df_clean.columns:
        if df_clean[col].dtype == "object":  # 文字列列
            df_clean[col] = df_clean[col].fillna("")
        else:  # 数値列
            df_clean[col] = df_clean[col].fillna(0)

    # xlsxwriterを使用して高品質出力
    try:
        output = io.BytesIO()
        # NaN/INFエラーを許可するオプションを追加
        with pd.ExcelWriter(
            output,
            engine="xlsxwriter",
            engine_kwargs={"options": {"nan_inf_to_errors": True}},
        ) as writer:
            df_clean.to_excel(
                writer, index=False, sheet_name=sheet_name, startrow=1, header=False
            )

            # xlsxwriterのworkbookとworksheetsオブジェクトを取得
            workbook = writer.book  # type: ignore
            worksheet = writer.sheets[sheet_name]  # type: ignore

            # フォント定義（游ゴシック、罫線なし）
            header_format = workbook.add_format(
                {  # type: ignore
                    "font_name": "游ゴシック",
                    "bold": True,
                    "bg_color": "#F2F2F2",
                }
            )

            cell_format = workbook.add_format({"font_name": "游ゴシック"})  # type: ignore

            unit_price_format = workbook.add_format(
                {  # type: ignore
                    "font_name": "游ゴシック",
                    "num_format": "#,##0.00",
                }
            )

            # ヘッダー書き込み
            for col_num, column_name in enumerate(df_clean.columns):
                worksheet.write(0, col_num, column_name, header_format)  # type: ignore

            # データ書き込み（単価だけフォーマットを分ける）
            for row_num in range(len(df_clean)):
                for col_num in range(len(df_clean.columns)):
                    col_name = df_clean.columns[col_num]
                    value = df_clean.iat[row_num, col_num]

                    if col_name == "単価":
                        worksheet.write(row_num + 1, col_num, value, unit_price_format)  # type: ignore
                    else:
                        worksheet.write(row_num + 1, col_num, value, cell_format)  # type: ignore

            # 列幅を個別に指定
            column_widths = {
                "大項目": 15,
                "中項目": 10,
                "合計正味重量": 10,
                "合計金額": 10,
                "単価": 7,
                "台数": 7,
            }

            for i, col_name in enumerate(df_clean.columns):
                width = column_widths.get(col_name, 20)
                worksheet.set_column(i, i, width)  # type: ignore

    except ImportError:
        # xlsxwriterがない場合はopenpyxlでフォールバック
        logger.warning("xlsxwriterが利用できません。openpyxlを使用します。")
        return simple_df_to_excel(df_clean, sheet_name)

    return output.getvalue()


def template_df_to_excel(
    df: pd.DataFrame, template_path: Union[str, Path], sheet_name: Optional[str] = None
) -> bytes:
    """
    テンプレートベースのDataFrame→Excel変換

    Args:
        df: 書き込むDataFrame（"セル"列と"値"列が必要）
        template_path: Excelテンプレートファイルのパス
        sheet_name: シート名（指定されている場合は変更）

    Returns:
        bytes: Excelファイルのバイナリデータ

    Raises:
        ImportError: openpyxlがインストールされていない場合
        FileNotFoundError: テンプレートファイルが見つからない場合
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.cell.cell import MergedCell
        from copy import copy
    except ImportError as e:
        raise ImportError(
            "テンプレート機能を使用するには openpyxl をインストールしてください"
        ) from e

    if isinstance(template_path, str):
        template_path = Path(template_path)

    if not template_path.exists():
        raise FileNotFoundError(
            f"テンプレートファイルが見つかりません: {template_path}"
        )

    # テンプレート読み込み
    wb = load_workbook(template_path)
    ws = wb.active

    if ws is None:
        raise ValueError("テンプレートファイルにアクティブシートが見つかりません")

    # DataFrameのデータをワークシートに書き込み
    for idx, row in df.iterrows():
        cell_ref = row.get("セル")
        value = safe_excel_value(row.get("値"))

        if pd.isna(cell_ref) or str(cell_ref).strip() in ["", "未設定"]:
            logger.info(f"空欄または未設定のセルはスキップされました。行 {idx}")
            continue

        try:
            # セル参照を正しく解析
            if ":" in str(cell_ref):
                logger.warning(f"範囲指定セル {cell_ref} はスキップしました")
                continue

            cell = ws[str(cell_ref)]

            # セル配列の場合、最初のセルを取得
            if isinstance(cell, tuple):
                cell = cell[0] if len(cell) > 0 else None

            if cell is None:
                logger.warning(f"セル {cell_ref} が取得できませんでした")
                continue

            if isinstance(cell, MergedCell):
                logger.warning(f"セル {cell_ref} は結合セルで書き込み不可。値: {value}")
                continue

            # 書式をdeep copyで保持
            try:
                original_font = copy(cell.font)
                original_fill = copy(cell.fill)
                original_border = copy(cell.border)
                original_format = cell.number_format

                # 値の上書き
                cell.value = value

                # 書式の復元
                cell.font = original_font
                cell.fill = original_fill
                cell.border = original_border
                cell.number_format = original_format
            except Exception as format_error:
                logger.warning(
                    f"セル {cell_ref} の書式保持に失敗、値のみ設定: {format_error}"
                )
                cell.value = value

        except Exception as e:
            logger.error(f"セル {cell_ref} 書き込み失敗: {e} / 値: {value}")

    # シート名変更（指定されている場合）
    if sheet_name:
        ws.title = sheet_name

    # メモリ出力
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def df_to_excel(
    df: pd.DataFrame,
    sheet_name: str = "データ",
    use_formatting: bool = True,
    template_path: Optional[Union[str, Path]] = None,
) -> bytes:
    """
    DataFrameをExcelファイル（bytes）に変換する統合関数

    Args:
        df: 出力するDataFrame
        sheet_name: シート名
        use_formatting: 日本語フォント等のフォーマットを適用するか
        template_path: テンプレートファイルのパス（指定時はテンプレート出力）

    Returns:
        bytes: Excelファイルのバイナリデータ
    """
    if template_path:
        return template_df_to_excel(df, template_path, sheet_name)
    elif use_formatting:
        return formatted_df_to_excel(df, sheet_name)
    else:
        return simple_df_to_excel(df, sheet_name)
